import datetime
import pandas as pd
from openpyxl import load_workbook

ABOUT = 0
DASHBOARD = 1
VALIDATION = 2
BIA = 3
BIA_DOCUMENTS = 4
PUBLISHER_FIELD = "Sample"


def have_documents(engine):
    documents = 0
    count = 0
    sql = "SELECT name FROM sqlite_master  WHERE type='table' and name like '%document%' and name not like 'ocds_%'"
    documentfTables = pd.read_sql_query(sql, engine)
    for index, row in documentfTables.iterrows():
        df = pd.read_sql_query("SELECT count() FROM " + row[0], engine)
        for index, row in df.iterrows():
            count = row[0]
        if count is not None:
            documents += count
    print("Documents: %d" % documents)
    return "No" if documents == 0 else "Yes", documents


def count_document_code(code, engine):
    documents = 0
    count = 0
    sql = "SELECT name FROM sqlite_master  WHERE type='table' and name like '%document%' and name not like 'ocds_%'"
    documentfTables = pd.read_sql_query(sql, engine)
    for index, row in documentfTables.iterrows():
        df = pd.read_sql_query("SELECT count() FROM " + row[0] + " where documentType = '%s'" % code, engine)
        for index, row in df.iterrows():
            count = row[0]
        if count is not None:
            documents += count
    return documents


def have_extensions(engine):
    extension = 0
    count = 0
    sql = "SELECT name FROM sqlite_master " \
          "WHERE type='table' and name not like 'ocds_%' and name not like '%publisher%' " \
          "and sql like '%extras%'"
    rows = pd.read_sql_query(sql, engine)
    for index, row in rows.iterrows():
        df = pd.read_sql_query("SELECT count() FROM " + row[0] + " where extras is not null", engine)
        for index, row in df.iterrows():
            count = row[0]
        if count is not None:
            extension += count
    print("Extensions: %d" % extension)
    return "No" if extension == 0 else "Yes", extension


def have_implementations(engine):
    documents = 0
    count = 0
    sql = "SELECT name FROM sqlite_master  WHERE type='table' and name like '%implementation%'"
    documentfTables = pd.read_sql_query(sql, engine)
    for index, row in documentfTables.iterrows():
        df = pd.read_sql_query("SELECT count() FROM " + row[0], engine)
        for index, row in df.iterrows():
            count = row[0]
        if count is not None:
            documents += count
    print("Implementations: %d" % documents)
    return "No" if documents == 0 else "Yes", documents


def have_buyers(engine):
    sql = "PRAGMA table_info(\"releases\")"
    count = 0
    columns = pd.read_sql_query(sql, engine)
    buyer_columns = "select count() from releases where "
    for index, row in columns.iterrows():
        if "buyer" in row[1]:
            buyer_columns += " " + row[1] + " is not null or "
    buyer_columns += " 1 = 0"
    df = pd.read_sql_query(buyer_columns, engine)
    for index, row in df.iterrows():
        count = row[0]
    if count is None:
        count = 0
    print("Buyers: %d" % count)
    return "No" if count == 0 else "Yes", count


def have_suppliers(engine):
    count = pd.read_sql_query("select count() from awards_suppliers", engine)
    for index, row in count.iterrows():
        count = row[0]
    if count is None:
        count = 0
    print("Suppliers: %d" % count)
    return "No" if count == 0 else "Yes", count


def replace_sample_name_aux(workbook, country, index, about_cells):
    ws = workbook.worksheets[index]
    for cell in about_cells:
        ws[cell] = ws[cell].value.replace(PUBLISHER_FIELD, country)


def generate_dashboard(workbook, fileNumber, schema, country):
    replace_sample_name_aux(workbook, country, DASHBOARD, ['B1', 'B26', 'B33', 'N26'])
    ws = workbook.worksheets[DASHBOARD]
    ws['N30'] = schema
    ws['P30'] = fileNumber


def generate_about(engine, workbook, country, metaData, fileNumber):
    replace_sample_name_aux(workbook, country, ABOUT, ['B1', 'I3', 'M3'])
    stages = ["plnning", "tender", "award", "contract", "implementation", "buyer", "suppliers", "documents",
              "extensions"]
    ws = workbook.worksheets[ABOUT]
    row = 5
    for stage in stages:
        ws["I%d" % row] = stage
        if stage == "buyer":
            h, c = have_buyers(engine)
        elif stage == "suppliers":
            h, c = have_suppliers(engine)
        elif stage == "documents":
            h, c = have_documents(engine)
        elif stage == "implementation":
            h, c = have_implementations(engine)
        elif stage == "extensions":
            h, c = have_extensions(engine)
        else:
            h, c = have_stage(engine, stage)
        ws["J%d" % row] = h
        ws["K%d" % row] = c
        row += 1
    meta = [{"Published": "publishedDate"}, {"Released": "publishedDate"},
            {"License": "license"}, {"Publication Policy": "publicationPolicy"},
            {"Publisher Name": "publisher.name"}, {"Publisher Scheme": "publisher.scheme"},
            {"Publisher UID": "publisher.uri"}]
    row = 5
    for m in meta:
        for key in m:
            ws["M%d" % row] = key
            ws["N%d" % row] = metaData.get(m[key])[0]
        row += 1
    ws["M%d" % row] = "Number of Files"
    ws["N%d" % row] = fileNumber


def generate_validation(engine, woorkbook):
    ws = woorkbook.worksheets[VALIDATION]
    errors = pd.read_sql_query("select * from publisher_validation", engine)
    excel_row = 2
    error_code = ""
    error_number = 0
    for index, row in errors.iterrows():
        pos = row[3].find("in")
        if row[3][pos:] != error_code:
            error_number += 1
            error_code = row[3][pos:]
        ws["A%d" % excel_row] = row[1]
        ws["B%d" % excel_row] = error_number
        ws["C%d" % excel_row] = row[3]
        excel_row += 1


def generate_bia(workbook, engine):
    ws = workbook.worksheets[BIA]
    excelRow = 2
    while ws['L%d' % excelRow].value is not None:
        field = ws['L%d' % excelRow].value.replace("/", "_")
        publisher_field = pd.read_sql_query("select fieldName, fillRate, tableName from publisher_bia "
                                            "where fieldName = '%s' and tableName  = 'releases'" % field, engine)
        if publisher_field.shape[0] == 0:
            if 'id' in field:
                publisher_field = pd.read_sql_query("select fieldName, fillRate, tableName from publisher_bia "
                                                    "where tableName || '_id' = '%s'" % field, engine)
            else:
                publisher_field = pd.read_sql_query("select fieldName, fillRate, tableName from publisher_bia "
                                                    "where tableName || '_'||fieldName = '%s'" % field, engine)

        if publisher_field.shape[0] == 0:
            # fillrate
            ws["D%d" % excelRow] = 0
            # publisher_fields
            ws["C%d" % excelRow] = ''
        else:
            for index, row in publisher_field.iterrows():
                ws["C%d" % excelRow] = row[0]
                ws["D%d" % excelRow] = row[1]
                ws["I%d" % excelRow] = row[2]
                ws["J%d" % excelRow] = row[0]
        excelRow += 1


def generate_bia_documents(workbook, engine):
    ws = workbook.worksheets[BIA_DOCUMENTS]
    excelRow = 2
    while ws['C%d' % excelRow].value is not None:
        code = ws['C%d' % excelRow].value
        ws['F%d' % excelRow] = count_document_code(code, engine)
        excelRow += 1


def create_dashboard(contry, engine, metaData, fileNumber, version, lang):
    now = datetime.datetime.now()
    fileName = "%s OCDS Evaluation [%s %d].xlsx" % (contry, now.strftime("%B"), now.year)
    wb2 = load_workbook('%s.xlsx' % lang)
    # About
    generate_about(engine, wb2, contry, metaData, fileNumber)
    # Dashboard
    generate_dashboard(wb2, fileNumber, version, contry)
    # Validation
    generate_validation(engine, wb2)
    # BIA
    generate_bia(wb2, engine)
    generate_bia_documents(wb2, engine)
    wb2.save(fileName)


def have_stage(engine, stage):
    dfTables = pd.read_sql_query("select * from releases where tag = '[\"" + stage + "\"]'", engine)
    if dfTables.shape[0] > 0:
        return "Yes", dfTables.shape[0]
    else:
        return "No", dfTables.shape[0]
